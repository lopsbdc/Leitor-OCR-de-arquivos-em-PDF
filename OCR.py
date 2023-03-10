import pdfplumber
from openpyxl import Workbook, load_workbook
import logging

planilha = load_workbook("Dados.xlsx") #planilha onde os dados lidos serão inseridos
aba = planilha.active #usar a primeira aba

# criando um arquivo LOG
logging.basicConfig(filename='Extrair_PDF.log', filemode='a', format='%(asctime)s - %(levelname)s - %(message)s')


# variaveis para criar um caminho de pasta + nome do arquivo PDF a ser lido
inicio = "Separar Arquivos/Separados/NewComp"
final = ".pdf"
meio = 'pag'
start = 1
total = 50 #total de arquivos a serem verificados

while start <= total:

    pag = 1
    try:
        try:
            while pag <= 900: #numero máximo de páginas do arquivo em PDF
                
                #pegar a ultima linha disponivel da planilha
                total_linhas = aba.max_row 
                
                #criando o caminho + nome do arquivo da página atual
                pdfname = inicio + str(start) + meio + str(pag) + final
                
                # iniciando a leitura
                with pdfplumber.open(pdfname) as pdf: 
                    linha = total_linhas + 1
                    pagina1 = pdf.pages[0]
                    result = pagina1.extract_text()
                    resultado = result.split("\n") #separando tudo que foi lido por linha

                    #verificadores do PDF, pois cada arquivo possui um padrão de leitura
                    nulo = resultado[15] 
                    tipo = resultado[2]
                    tipo4 = resultado[0]

                    if tipo4 == 'Generated by PDFKit.NET Evaluation':

                        # Onde ficam as informações no PDF de tipo 4
                        razao1 = resultado[4]
                        data1 = resultado[9]
                        valor1 = resultado[13]
                        autenticacao1 = resultado[14]

                        # Onde ficam as demais informações no PDF de tipo 4
                        razao0 = resultado[18]
                        data0 = resultado[23]
                        valor0 = resultado[27]
                        autenticacao0 = resultado[28]

                        # Tratamento dos dados lidos da primeira parte
                        razao = razao1.replace("Nome do Cliente: ", "")
                        data = data1.replace("Data de Pagamento: ", "")
                        valor = valor1.replace("Valor de Pagamento: ", "")
                        autenticacao = autenticacao1.replace("Autenticação Eletrônica: ", "")
                        nomepdf = pdfname.replace("Separar Arquivos/Separados/", "")

                        # Tratamento dos dados lidos da segunda parte
                        razao2 = razao0.replace("Nome do Cliente: ", "")
                        data2 = data0.replace("Data de Pagamento: ", "")
                        valor2 = valor0.replace("Valor de Pagamento: ", "")
                        autenticacao2 = autenticacao0.replace("Autenticação Eletrônica: ", "")
                        nomepdf2 = pdfname.replace("Separar Arquivos/Separados/", "")

                        # Inserindo os dados da primeira parte na planilha
                        aba.cell(row=linha, column=1).value = razao
                        aba.cell(row=linha, column=2).value = data
                        aba.cell(row=linha, column=3).value = valor
                        aba.cell(row=linha, column=4).value = autenticacao
                        aba.cell(row=linha, column=5).value = nomepdf

                        # Inserindo os dados da segunda parte na planilha
                        aba.cell(row=linha + 1, column=1).value = razao2
                        aba.cell(row=linha + 1, column=2).value = data2
                        aba.cell(row=linha + 1, column=3).value = valor2
                        aba.cell(row=linha + 1, column=4).value = autenticacao2
                        aba.cell(row=linha + 1, column=5).value = nomepdf2

                        logging.warning("PDF tipo 4 " + nomepdf + " concluido com sucesso")
                    
                    # PDF de tipo 1
                    elif tipo == "MAGAZINE LUIZA S/A":
                      
                        #leitura dos dados
                        razao3 = resultado[9]
                        data3 = resultado[17]
                        valor3 = resultado[21]
                        autenticacao3 = resultado[23]
                        
                        #tratamento dos dados
                        razao = razao3.replace("Beneficiário: ", "")
                        data = data3.replace("Data de Pagamento: ", "")
                        valor = valor3.replace("Valor Cobrado: ", "")
                        autenticacao = autenticacao3.replace("Autenticação: ", "")
                        nomepdf = pdfname.replace("Separar Arquivos/Separados/", "")
                        
                        #inserção dos dados na planilha
                        aba.cell(row=linha, column=1).value = razao
                        aba.cell(row=linha, column=2).value = data
                        aba.cell(row=linha, column=3).value = valor
                        aba.cell(row=linha, column=4).value = autenticacao
                        aba.cell(row=linha, column=5).value = nomepdf

                        logging.warning("PDF tipo 1 " + nomepdf + " concluido com sucesso")

                    else:
                        
                        # PDF de tipo 2
                        if nulo != "null":

                            # Onde ficam as informações no PDF de tipo 2
                            razao1 = resultado[3]
                            data1 = resultado[8]
                            valor1 = resultado[12]
                            autenticacao1 = resultado[13]

                            # Onde ficam as demais informações no PDF de tipo 2
                            razao0 = resultado[17]
                            data0 = resultado[22]
                            valor0 = resultado[26]
                            autenticacao0 = resultado[27]

                            # Tratamento dos dados lidos da primeira parte
                            razao = razao1.replace("Nome do Cliente: ", "")
                            data = data1.replace("Data de Pagamento: ", "")
                            valor = valor1.replace("Valor de Pagamento: ", "")
                            autenticacao = autenticacao1.replace("Autenticação Eletrônica: ", "")
                            nomepdf = pdfname.replace("Separar Arquivos/Separados/", "")

                            # Tratamento dos dados lidos da segunda parte
                            razao2 = razao0.replace("Nome do Cliente: ", "")
                            data2 = data0.replace("Data de Pagamento: ", "")
                            valor2 = valor0.replace("Valor de Pagamento: ", "")
                            autenticacao2 = autenticacao0.replace("Autenticação Eletrônica: ", "")
                            nomepdf2 = pdfname.replace("Separar Arquivos/Separados/", "")

                            #Inserindo os dados da primeira parte na planilha
                            aba.cell(row=linha, column=1).value = razao
                            aba.cell(row=linha, column=2).value = data
                            aba.cell(row=linha, column=3).value = valor
                            aba.cell(row=linha, column=4).value = autenticacao
                            aba.cell(row=linha, column=5).value = nomepdf

                            #Inserindo os dados da segunda parte na planilha
                            aba.cell(row=linha+1, column=1).value = razao2
                            aba.cell(row=linha+1, column=2).value = data2
                            aba.cell(row=linha+1, column=3).value = valor2
                            aba.cell(row=linha+1, column=4).value = autenticacao2
                            aba.cell(row=linha+1, column=5).value = nomepdf2

                            logging.warning("PDF tipo 2 " + nomepdf + " concluido com sucesso")
                        
                        # PDF de tipo 3
                        elif nulo == "null":

                            # Onde ficam as informações no PDF de tipo 3
                            razao1 = resultado[3]
                            data1 = resultado[8]
                            valor1 = resultado[12]
                            autenticacao1 = resultado[13]

                            # Tratamento dos dados lidos
                            razao = razao1.replace("Nome do Cliente: ", "")
                            data = data1.replace("Data de Pagamento: ", "")
                            valor = valor1.replace("Valor de Pagamento: ", "")
                            autenticacao = autenticacao1.replace("Autenticação Eletrônica: ", "")
                            nomepdf = pdfname.replace("Separar Arquivos/Separados/", "")

                            # Inserção dos dados na planilha
                            aba.cell(row=linha, column=1).value = razao
                            aba.cell(row=linha, column=2).value = data
                            aba.cell(row=linha, column=3).value = valor
                            aba.cell(row=linha, column=4).value = autenticacao
                            aba.cell(row=linha, column=5).value = nomepdf

                            logging.warning("PDF tipo 3 " + nomepdf + " concluido com sucesso")

                        else:
                            logging.warning("Não identificado o tipo de comprovante, para o pdf " + pdfname)
                print('pagina ' + str(pag) + ' foi finalizada')
                pag = pag + 1

        except:
          #otimizador de tempo, caso o PDF possua poucas páginas
            if pag >= 20:
                pag = pag + 900
                logging.warning("não encontrado nenhuma pagina após o pdf " + nomepdf)
            else:
                pag = pag + 1
    except:
        logging.warning('erro desconhecido')


    start = start + 1
    
#salvar a planilha somente no final do script, evitando corromper a planilha, bem como aumentar a produtividade e reduzir o gasto de recursos
planilha.save('Dados.xlsx')


